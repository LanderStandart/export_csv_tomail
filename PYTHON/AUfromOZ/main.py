import json

from  engine import Archiv,Db,read_ini,my_log,clear_export_path,Mysql
import openpyxl as XLS
import datetime,os
import locale,sys
from tqdm import tqdm
import smtplib
from email import encoders
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart

logger = my_log.get_logger(__name__)

logger.info('Init...')
locale.setlocale(
    category=locale.LC_ALL,
    locale="Russian")
section = 'BASE_CONF'
host = read_ini(section, 'HOST')
database = read_ini(section, 'PATH')
DB = Db(host,database)
path = read_ini(section, 'PATH_EXPORT')
MYSQL = Mysql()
typeprotocol = json.loads(read_ini(section,'typeprotocol'))
textprotocol=json.loads(read_ini(section,'textprotocol'))
listpost =json.loads(read_ini(section,'listpost'))
toemail =read_ini(section,'toemail').split(',')

def CheckDayofWeek():
    weekday = datetime.datetime.now().strftime('%w')
    if weekday  in ('0','6'):
        exit('Выходной')
    if weekday == '1':
        data_start = datetime.datetime.now()-datetime.timedelta(days=3)
    else:
        data_start = datetime.datetime.now() - datetime.timedelta(days=1)
    return (data_start)


#конец инициализации
def CreateProtocol(firm,typeP,index,postid):
    class_tovar = '1' if (int(key) % 2 >0) else '0'
    wb = XLS.load_workbook('head.xlsx')
    ws = wb.active
    ws1 = wb.create_sheet("Сумма по Аптекам")

    data_start = CheckDayofWeek()
    print(data_start)
    task = 'zakaz'
    val ={'datestart':data_start.strftime('%Y-%m-%d'),'client_name':firm,'postid':postid,'class':class_tovar}
    logger.info(val)
    data = MYSQL.get_from_base(__name__, task, val)

    pbar = tqdm(total=len(data))
    for row in data:
        val = {'id':row[1]}
        sql = 'getName'
        name =DB.get_from_base(__name__,sql,val)
        tru_name = ''.join(str(name[0])).replace("['","").replace("']","")
        row[1] = tru_name
        ws.append(row)
        pbar.update(1)

    task = 'grzakaz'
    val = {'datestart': data_start.strftime('%Y-%m-%d'), 'client_name': firm, 'postid': postid,'class':class_tovar}
    data1 = MYSQL.get_from_base(__name__, task, val)
    pbar = tqdm(total=len(data1))
    for row in data1:
        ws1.append(row)
        pbar.update(1)
    #
    EditHeader(ws,firm,typeP,index)
    ws1.column_dimensions['A'].width=30
    ws1.column_dimensions['B'].width = 70
    ws1.column_dimensions['C'].width = 10

    ws.cell(23 + len(data), 5, value='ИТОГО')
    if len(data)>0:
        ws.cell(23 + len(data),6,value=f'=SUM(F23:F{22 + len(data)})')
    else:
        ws.cell(23 + len(data), 6, value=f'=0')

    type_protokol = json.loads(read_ini(section,'index_p'))
    print(class_tovar)
    print(type_protokol[typeP])
    f_name = f'{path}Протокол ЗЦ в серии закупок на закупку {type_protokol[typeP]} {firm} - {data_start.strftime("%Y-%m-%d")}'
    wb.save(f'{f_name}.xls')
    MailTo(f_name)


def EditHeader(ws,firm,typeP,index):
    current_date = datetime.datetime.now().strftime('"%d" %B %Y')
    #Заполняем Юр Лицо.
    ws.cell(5, 6, current_date)
    proto = ws['A3'].value
# если текст большой увеличваем высоту строки
    rowhigh = ws.row_dimensions[3].height if index not in '2' else ws.row_dimensions[3].height * 3
    ws.row_dimensions[3].height = rowhigh

    ws['A3'].value = proto.replace('"11"', f'"{firm}"').replace('textprotocol',textprotocol[index])
    proto = ws['A8'].value
    ws['A8'].value = proto.replace('"11"', f'"{firm.upper()}"')
    ws['A2'].value = typeprotocol[str(typeP)]
    return ws

def MailTo(filename):
    Archiv(filename, 'xls').zip_File()
    basename = os.path.basename(f'{filename}.zip')
    from_addr ="fs@standart-n.ru"
    file_to_attach = f'{filename}.zip'
    to_emails = toemail
    msg = MIMEMultipart()
    msg["From"] = "fs@standart-n.ru"
    msg["Subject"] = basename
    msg["To"] = ", ".join(to_emails)
    msg.attach(MIMEText('text'))

    try:
        with open(file_to_attach, "rb") as fil:
            part = MIMEApplication(
                fil.read(),
                Name=basename
            )
        # After the file is closed
        part['Content-Disposition'] = 'attachment; filename="%s"' % basename
        msg.attach(part)
    except IOError:
        print(f"Ошибка при открытии файла вложения {file_to_attach}")

    try:
        smtp = smtplib.SMTP('192.168.67.235')
     #   smtp.starttls()
        smtp.ehlo()
        smtp.login('fs', 'teric0ls1')
        smtp.sendmail(from_addr, to_emails, msg.as_string())
    except smtplib.SMTPException as err:
        print('Что - то пошло не так...')
        raise err
    finally:
        smtp.quit()






section = 'BASE_CONF'
clients = read_ini(section,'clients').split(',')

for name in clients:

      for key in typeprotocol.keys():
          index = '1' if (int(key) % 2 >0) else '2'
          CreateProtocol(name,key,index,listpost[f'{name} {key}'])

